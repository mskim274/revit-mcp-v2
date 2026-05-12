"""Build an xlsx of all RC/SRC column family types from the active Revit document.

Data was queried via revit-mcp on 2026-05-06 and is embedded as a constant below.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import re

OUTPUT = r"C:\Users\and\Desktop\revit-mcp-v2\Y1P1_기둥타입_정리_2026-05-06.xlsx"

# (family, type_id, type_name)
DATA = [
    # ─── RC-Column_포디움(A구간) — 129 ─────────────────────────────────
    ("RC-Column_포디움(A구간)", 476159, "Column_RC, AC0, 40MPa, 700x700, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476169, "Column_RC, AC0A-1, 40MPa, 800x800, 2F"),
    ("RC-Column_포디움(A구간)", 476167, "Column_RC, AC0A-1, 40MPa, 800x800, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476163, "Column_RC, AC0A, 40MPa, 800x800, B2F-2F"),
    ("RC-Column_포디움(A구간)", 476165, "Column_RC, AC0B, SK_MATE, 800x800, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476173, "Column_RC, AC0C, 40MPa, 1000x1200, 2F"),
    ("RC-Column_포디움(A구간)", 476171, "Column_RC, AC0C, 40MPa, 1000x1200, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476177, "Column_RC, AC0D, 40MPa, 900x1400, 2F"),
    ("RC-Column_포디움(A구간)", 476175, "Column_RC, AC0D, 40MPa, 900x1400, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476187, "Column_RC, AC1-1, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476185, "Column_RC, AC1-1, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476183, "Column_RC, AC1, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476181, "Column_RC, AC1, 40MPa, 1100x1100, 1F-2F"),
    ("RC-Column_포디움(A구간)", 476179, "Column_RC, AC1, 40MPa, 1200x1200, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 5223481, "Column_RC, AC10, 35MPa, 1100x1600, 2F"),
    ("RC-Column_포디움(A구간)", 476251, "Column_RC, AC10, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(A구간)", 5223479, "Column_RC, AC10, 40MPa, 1100x1600, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476375, "Column_RC, AC100, 40MPa, 1000x1000, B2F-2F"),
    ("RC-Column_포디움(A구간)", 476377, "Column_RC, AC100A, SK_MATE, 1000x1000, B2F-3F"),
    ("RC-Column_포디움(A구간)", 476379, "Column_RC, AC100B, SK_MATE, 1000x1000, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476383, "Column_RC, AC101, 35MPa, 1700x1700, 5F"),
    ("RC-Column_포디움(A구간)", 476387, "Column_RC, AC102, 35MPa, 1700x1700, 5F"),
    ("RC-Column_포디움(A구간)", 476391, "Column_RC, AC103, 35MPa, 1700x1700, 5F"),
    ("RC-Column_포디움(A구간)", 476397, "Column_RC, AC104-1, 49MPa, 1700x1700, 4F-5F"),
    ("RC-Column_포디움(A구간)", 4969744, "Column_RC, AC104, 49MPa, 1700x1700, 4F-5F"),
    ("RC-Column_포디움(A구간)", 476403, "Column_RC, AC104A-1, SK_MATE, 1700x1700, B2F-5F"),
    ("RC-Column_포디움(A구간)", 476399, "Column_RC, AC104A, SK_MATE, 1700x1700, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476407, "Column_RC, AC105, 35MPa, 1400x1400, 4F"),
    ("RC-Column_포디움(A구간)", 476417, "Column_RC, AC105A-1, 35MPa, 1100x1100, 5F"),
    ("RC-Column_포디움(A구간)", 476415, "Column_RC, AC105A-1, SK_MATE, 1100x1100, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476409, "Column_RC, AC105A, SK_MATE, 1100x1100, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476413, "Column_RC, AC105B, SK_MATE, 1400x1400, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476425, "Column_RC, AC106-1, SK_MATE, 1700x1700, B2F-5F"),
    ("RC-Column_포디움(A구간)", 476421, "Column_RC, AC106, SK_MATE, 1700x1700, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476431, "Column_RC, AC108, SK_MATE, 1700x1700, B2F-4F"),
    ("RC-Column_포디움(A구간)", 5318717, "Column_RC, AC109, 49MPa, 1100x1100, 5F"),
    ("RC-Column_포디움(A구간)", 476255, "Column_RC, AC10A, 40MPa, 1100x1600, 2F"),
    ("RC-Column_포디움(A구간)", 476253, "Column_RC, AC10A, 40MPa, 1100x1600, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476259, "Column_RC, AC11, 40MPa, 800x800, 1F"),
    ("RC-Column_포디움(A구간)", 476257, "Column_RC, AC11, 40MPa, 800x800, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476263, "Column_RC, AC11A, 40MPa, 800x800, 1F"),
    ("RC-Column_포디움(A구간)", 476261, "Column_RC, AC11A, 40MPa, 800x800, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476265, "Column_RC, AC12, SK_MATE, 800x800, 2F-3F"),
    ("RC-Column_포디움(A구간)", 476267, "Column_RC, AC13, 40MPa, 800x800, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476269, "Column_RC, AC13, SK_MATE, 800x800, 1F-3F"),
    ("RC-Column_포디움(A구간)", 476273, "Column_RC, AC13A, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(A구간)", 476271, "Column_RC, AC13A, 40MPa, 800x800, B2F-2F"),
    ("RC-Column_포디움(A구간)", 476275, "Column_RC, AC15, 40MPa, 800x800, B2F-2F"),
    ("RC-Column_포디움(A구간)", 476277, "Column_RC, AC15A, 40MPa, 600x600, B2F-2F"),
    ("RC-Column_포디움(A구간)", 476279, "Column_RC, AC15B, 40MPa, 800x800, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476287, "Column_RC, AC16-1, 40MPa, 800x800, 2F"),
    ("RC-Column_포디움(A구간)", 476285, "Column_RC, AC16, 40MPa, 800x800, 1F-2F"),
    ("RC-Column_포디움(A구간)", 476283, "Column_RC, AC16, 40MPa, 800x800, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476293, "Column_RC, AC18-1, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 4936884, "Column_RC, AC18, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476289, "Column_RC, AC18, 40MPa, 1100x1100, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476191, "Column_RC, AC1A, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 4938073, "Column_RC, AC1A, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476197, "Column_RC, AC1B, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476195, "Column_RC, AC1B, 40MPa, 1100x1100, 1F-2F"),
    ("RC-Column_포디움(A구간)", 476193, "Column_RC, AC1B, 40MPa, 1200x1200, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 5223477, "Column_RC, AC2-2, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 5223475, "Column_RC, AC2-2, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476201, "Column_RC, AC2, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476199, "Column_RC, AC2, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476299, "Column_RC, AC21-1, SK_MATE, 900x900, 2F-4F"),
    ("RC-Column_포디움(A구간)", 476295, "Column_RC, AC21, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476297, "Column_RC, AC21, SK_MATE, 900x900, 1F-4F"),
    ("RC-Column_포디움(A구간)", 476305, "Column_RC, AC21A, 35MPa, 900x900, 4F"),
    ("RC-Column_포디움(A구간)", 476301, "Column_RC, AC21A, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476303, "Column_RC, AC21A, SK_MATE, 900x900, 1F-3F"),
    ("RC-Column_포디움(A구간)", 476321, "Column_RC, AC22-1, 35MPa, 900x900, 4F"),
    ("RC-Column_포디움(A구간)", 476319, "Column_RC, AC22-1, SK_MATE, 900x900, 2F-3F"),
    ("RC-Column_포디움(A구간)", 476307, "Column_RC, AC22, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476309, "Column_RC, AC22, SK_MATE, 900x900, 1F-4F"),
    ("RC-Column_포디움(A구간)", 2561814, "Column_RC, AC22A-1, 35MPa, 900x900, 2F-3F"),
    ("RC-Column_포디움(A구간)", 476317, "Column_RC, AC22A, 35MPa, 900x900, 4F"),
    ("RC-Column_포디움(A구간)", 476313, "Column_RC, AC22A, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476315, "Column_RC, AC22A, SK_MATE, 900x900, 1F-3F"),
    ("RC-Column_포디움(A구간)", 476327, "Column_RC, AC23, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476325, "Column_RC, AC23, 40MPa, 1000x1000, 1F-2F"),
    ("RC-Column_포디움(A구간)", 476323, "Column_RC, AC23, 40MPa, 1200x1000, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476333, "Column_RC, AC23A, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476331, "Column_RC, AC23A, 40MPa, 1000x1000, 2F"),
    ("RC-Column_포디움(A구간)", 476329, "Column_RC, AC23A, 40MPa, 1100x1100, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476335, "Column_RC, AC23B, SK_MATE, 900x900, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476337, "Column_RC, AC25, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476339, "Column_RC, AC25, SK_MATE, 900x900, 1F-4F"),
    ("RC-Column_포디움(A구간)", 476345, "Column_RC, AC25A, 35MPa, 900x900, 4F"),
    ("RC-Column_포디움(A구간)", 476341, "Column_RC, AC25A, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476343, "Column_RC, AC25A, SK_MATE, 900x900, 1F-3F"),
    ("RC-Column_포디움(A구간)", 476355, "Column_RC, AC28-1, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476353, "Column_RC, AC28-1, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476351, "Column_RC, AC28, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476349, "Column_RC, AC28, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476347, "Column_RC, AC28, 40MPa, 1100x1100, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476361, "Column_RC, AC28A, 35MPa, 1000x1000, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476359, "Column_RC, AC28A, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476357, "Column_RC, AC28A, 40MPa, 1100x1100, B2F-1F"),
    ("RC-Column_포디움(A구간)", 476365, "Column_RC, AC29, 35MPa, 900x900, 4F"),
    ("RC-Column_포디움(A구간)", 476363, "Column_RC, AC29, SK_MATE, 900x900, B2F-3F"),
    ("RC-Column_포디움(A구간)", 476211, "Column_RC, AC3-1, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476209, "Column_RC, AC3-1, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476207, "Column_RC, AC3, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476205, "Column_RC, AC3, 40MPa, 1100x1100, 1F-2F"),
    ("RC-Column_포디움(A구간)", 476203, "Column_RC, AC3, 40MPa, 1200x1200, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476367, "Column_RC, AC30, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(A구간)", 476369, "Column_RC, AC30, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(A구간)", 476371, "Column_RC, AC30A, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(A구간)", 476373, "Column_RC, AC30A, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(A구간)", 476215, "Column_RC, AC3A, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476213, "Column_RC, AC3A, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476219, "Column_RC, AC6-1, SK_MATE, 800x800, 2F-4F"),
    ("RC-Column_포디움(A구간)", 476217, "Column_RC, AC6, SK_MATE, 800x800, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476221, "Column_RC, AC6A, SK_MATE, 800x800, B2F-4F"),
    ("RC-Column_포디움(A구간)", 476231, "Column_RC, AC7-1, 35MPa, 900x900, 4F"),
    ("RC-Column_포디움(A구간)", 476229, "Column_RC, AC7-1, SK_MATE, 900x900, 2F-3F"),
    ("RC-Column_포디움(A구간)", 476227, "Column_RC, AC7, 35MPa, 900x900, 4F"),
    ("RC-Column_포디움(A구간)", 476223, "Column_RC, AC7, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476225, "Column_RC, AC7, SK_MATE, 900x900, 1F-3F"),
    ("RC-Column_포디움(A구간)", 476241, "Column_RC, AC8-1, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476239, "Column_RC, AC8-1, 40MPa, 1100x1100, 2F"),
    ("RC-Column_포디움(A구간)", 476237, "Column_RC, AC8, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(A구간)", 476235, "Column_RC, AC8, 40MPa, 1100x1100, 1F-2F"),
    ("RC-Column_포디움(A구간)", 476233, "Column_RC, AC8, 40MPa, 1100x1100, B2F-B1F"),
    ("RC-Column_포디움(A구간)", 476245, "Column_RC, AC9, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(A구간)", 2959644, "Column_RC, AC9, 35MPa, 1200x1200, 5F-6F"),
    ("RC-Column_포디움(A구간)", 476243, "Column_RC, AC9, 40MPa, 1200x1200, 2F"),
    ("RC-Column_포디움(A구간)", 4605826, "Column_RC, NA, SK_MATE, 400x400, 도면누락"),

    # ─── RC-Column_포디움(B,C구간) — 264 ───────────────────────────────
    ("RC-Column_포디움(B,C구간)", 486532, "Column_RC, BC1, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486530, "Column_RC, BC1, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486528, "Column_RC, BC1, 40MPa, 1400x1400, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486584, "Column_RC, BC11, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486582, "Column_RC, BC11, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486590, "Column_RC, BC11A, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486588, "Column_RC, BC11A, 40MPa, 1300x1300, 2F"),
    ("RC-Column_포디움(B,C구간)", 486596, "Column_RC, BC12, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486594, "Column_RC, BC12, 40MPa, 1300x1300, 1F"),
    ("RC-Column_포디움(B,C구간)", 486602, "Column_RC, BC12A, 35MPa, 1000x1000, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486600, "Column_RC, BC12A, 40MPa, 1400x1400, 1F"),
    ("RC-Column_포디움(B,C구간)", 486606, "Column_RC, BC12B, 35MPa, 1000x1000, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486610, "Column_RC, BC13, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 973244, "Column_RC, BC13, 40MPa, 1300x1300, 1F"),
    ("RC-Column_포디움(B,C구간)", 486540, "Column_RC, BC1A, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486538, "Column_RC, BC1A, 40MPa, 1400x1400, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486536, "Column_RC, BC1A, 40MPa, 1500x1500, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486558, "Column_RC, BC2A-1, 35MPa, 1200x1200, 5F"),
    ("RC-Column_포디움(B,C구간)", 486556, "Column_RC, BC2A-1, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486554, "Column_RC, BC2A-1, 40MPa, 1400x1400, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486552, "Column_RC, BC2A-1, 40MPa, 1500x1500, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486548, "Column_RC, BC2A, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486546, "Column_RC, BC2A, 40MPa, 1400x1400, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486544, "Column_RC, BC2A, 40MPa, 1500x1500, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486566, "Column_RC, BC3, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486564, "Column_RC, BC3, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486562, "Column_RC, BC3, 40MPa, 1400x1400, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486578, "Column_RC, BC3A, 35MPa, 1100x1100, 5F"),
    ("RC-Column_포디움(B,C구간)", 486576, "Column_RC, BC3A, 35MPa, 1200x1200, 4F"),
    ("RC-Column_포디움(B,C구간)", 486574, "Column_RC, BC3A, 35MPa, 2800x1200, 3F"),
    ("RC-Column_포디움(B,C구간)", 486572, "Column_RC, BC3A, 40MPa, 2800x1200, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486570, "Column_RC, BC3A, 40MPa, 2800x1200, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486218, "Column_RC, C0, 40MPa, 700x700, B2F-1F"),
    ("RC-Column_포디움(B,C구간)", 486220, "Column_RC, C0A, 40MPa, 800x800, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486224, "Column_RC, C0B-2, SK_MATE, 800x800, B2F-4F"),
    ("RC-Column_포디움(B,C구간)", 486244, "Column_RC, C1, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486240, "Column_RC, C1, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486242, "Column_RC, C1, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486298, "Column_RC, C10-2, 40MPa, 600x600, B2F-2F"),
    ("RC-Column_포디움(B,C구간)", 486296, "Column_RC, C10, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486464, "Column_RC, C101, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486462, "Column_RC, C101, SK_MATE, 2800x1200, B2F-3F"),
    ("RC-Column_포디움(B,C구간)", 486302, "Column_RC, C11, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486300, "Column_RC, C11, 40MPa, 900x900, 2F"),
    ("RC-Column_포디움(B,C구간)", 486468, "Column_RC, C111, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486466, "Column_RC, C111, SK_MATE, 2000x1400, B1F-3F"),
    ("RC-Column_포디움(B,C구간)", 486472, "Column_RC, C111A, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486470, "Column_RC, C111A, SK_MATE, 2000x1400, B1F-3F"),
    ("RC-Column_포디움(B,C구간)", 486476, "Column_RC, C112, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486474, "Column_RC, C112, SK_MATE, 2000x1400, 1F-3F"),
    ("RC-Column_포디움(B,C구간)", 486480, "Column_RC, C113, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486478, "Column_RC, C113, SK_MATE, 2000x1400, 2F-3F"),
    ("RC-Column_포디움(B,C구간)", 486312, "Column_RC, C11B, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486310, "Column_RC, C11B, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486320, "Column_RC, C11C-1, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486318, "Column_RC, C11C, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486314, "Column_RC, C11C, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 486316, "Column_RC, C11C, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486482, "Column_RC, C131, 35MPa, 2000x1400, 3F"),
    ("RC-Column_포디움(B,C구간)", 486484, "Column_RC, C131, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486328, "Column_RC, C13A, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 486334, "Column_RC, C15, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486342, "Column_RC, C15A, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486340, "Column_RC, C15A, 40MPa, 800x800, 1F"),
    ("RC-Column_포디움(B,C구간)", 486344, "Column_RC, C15B, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486346, "Column_RC, C16, 35MPa, 800x800, 3F-5F"),
    ("RC-Column_포디움(B,C구간)", 486250, "Column_RC, C1A-1, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486246, "Column_RC, C1A-1, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486248, "Column_RC, C1A-1, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 4984371, "Column_RC, C1A, 35MPa, 800x800, 3F-5F"),
    ("RC-Column_포디움(B,C구간)", 4888000, "Column_RC, C1A, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 4942773, "Column_RC, C1A, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486256, "Column_RC, C2, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 486252, "Column_RC, C2, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486254, "Column_RC, C2, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486358, "Column_RC, C21-1, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486354, "Column_RC, C21-1, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486356, "Column_RC, C21-1, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486352, "Column_RC, C21, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486348, "Column_RC, C21, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486350, "Column_RC, C21, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486362, "Column_RC, C21A, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486360, "Column_RC, C21A, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486366, "Column_RC, C22-1, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 4888009, "Column_RC, C22-2, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486364, "Column_RC, C22, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486370, "Column_RC, C22A, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486376, "Column_RC, C22B, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 486372, "Column_RC, C22B, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486374, "Column_RC, C22B, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486382, "Column_RC, C22C-1, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486384, "Column_RC, C22C-1, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486390, "Column_RC, C23, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486386, "Column_RC, C23, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486388, "Column_RC, C23, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486400, "Column_RC, C23C, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486398, "Column_RC, C23C, 40MPa, 1400x1400, 1F"),
    ("RC-Column_포디움(B,C구간)", 486396, "Column_RC, C23C, 40MPa, 1400x1400, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486406, "Column_RC, C24, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486402, "Column_RC, C24, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486404, "Column_RC, C24, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486412, "Column_RC, C24A-1, 40MPa, 900x900, 2F"),
    ("RC-Column_포디움(B,C구간)", 486414, "Column_RC, C24B, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486416, "Column_RC, C24B, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486422, "Column_RC, C25, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 486420, "Column_RC, C25, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486418, "Column_RC, C25, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486264, "Column_RC, C3, 40MPa, 800x800, 1F"),
    ("RC-Column_포디움(B,C구간)", 486262, "Column_RC, C3, 40MPa, 800x800, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486430, "Column_RC, C31-1, 35MPa, 800x800, 3F-5F"),
    ("RC-Column_포디움(B,C구간)", 486428, "Column_RC, C31-1, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486426, "Column_RC, C31, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486424, "Column_RC, C31, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486434, "Column_RC, C31A, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486440, "Column_RC, C32, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 486438, "Column_RC, C32, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486446, "Column_RC, C32B-1, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 486450, "Column_RC, C32C, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486448, "Column_RC, C32C, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 486452, "Column_RC, C35, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 486454, "Column_RC, C36, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 5345263, "Column_RC, C36A-1, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486456, "Column_RC, C36A, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486460, "Column_RC, C36B, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486266, "Column_RC, C3A, 40MPa, 800x800, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486288, "Column_RC, C4-1, 40MPa, 900 x 900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486286, "Column_RC, C4-1, 40MPa, 900 x900 , B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486294, "Column_RC, C4-2, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486272, "Column_RC, C4, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486270, "Column_RC, C4, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486274, "Column_RC, C4, SK_MATE, 800x800, 2F-4F"),
    ("RC-Column_포디움(B,C구간)", 486276, "Column_RC, C4A, 40MPa, 1000x900, B2F-1F"),
    ("RC-Column_포디움(B,C구간)", 486278, "Column_RC, C4A, SK_MATE, 800x800, 2F-4F"),
    ("RC-Column_포디움(B,C구간)", 486284, "Column_RC, C4B, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486282, "Column_RC, C4B, 40MPa, 800x800, 1F"),
    ("RC-Column_포디움(B,C구간)", 486280, "Column_RC, C4B, 40MPa, 800x800, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486648, "Column_RC, CC11, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 4952619, "Column_RC, CC11, 40MPa, 1200x1200, 2F"),
    ("RC-Column_포디움(B,C구간)", 486652, "Column_RC, CC11A, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486650, "Column_RC, CC11A, 40MPa, 900x900, 2F"),
    ("RC-Column_포디움(B,C구간)", 5357487, "Column_RC, CC11B, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486664, "Column_RC, CC12A-1, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 2670848, "Column_RC, CC12A-2, 35MPa, 800x800, 4F"),
    ("RC-Column_포디움(B,C구간)", 486666, "Column_RC, CC12A-3, 35MPa, 800x800, 3F-5F"),
    ("RC-Column_포디움(B,C구간)", 486656, "Column_RC, CC12A, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486660, "Column_RC, CC12B, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 5331968, "Column_RC, CC13-1, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486672, "Column_RC, CC13, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486670, "Column_RC, CC13, 40MPa, 900x900, 2F"),
    ("RC-Column_포디움(B,C구간)", 486686, "Column_RC, CC15-1, 35MPa, 800x800, 3F-5F"),
    ("RC-Column_포디움(B,C구간)", 486676, "Column_RC, CC15, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486680, "Column_RC, CC15A, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486684, "Column_RC, CC15B, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 486614, "Column_RC, CC2A, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486622, "Column_RC, CC3, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486618, "Column_RC, CC3, 40MPa, 1000x1000, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486620, "Column_RC, CC3, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 486644, "Column_RC, CC5-1, 35MPa, 800x800, 3F-5F"),
    ("RC-Column_포디움(B,C구간)", 486642, "Column_RC, CC5-1, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486640, "Column_RC, CC5-1, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486630, "Column_RC, CC5, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486628, "Column_RC, CC5, 40MPa, 800x800, 1F"),
    ("RC-Column_포디움(B,C구간)", 486626, "Column_RC, CC5, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486638, "Column_RC, CC5A, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 486636, "Column_RC, CC5A, 40MPa, 900x900, 1F"),
    ("RC-Column_포디움(B,C구간)", 486634, "Column_RC, CC5A, 40MPa, 900x900, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 633105, "Column_RC, DC1, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633095, "Column_RC, DC1, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633085, "Column_RC, DC1, 40MPa, 1300x1300, B1F"),
    ("RC-Column_포디움(B,C구간)", 633580, "Column_RC, DC11, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633560, "Column_RC, DC11, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 633570, "Column_RC, DC11, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633440, "Column_RC, DC11A, 35MPa, 900x900, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633428, "Column_RC, DC11A, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 633438, "Column_RC, DC11A, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633462, "Column_RC, DC12, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633450, "Column_RC, DC12, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 633460, "Column_RC, DC12, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633602, "Column_RC, DC13, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633590, "Column_RC, DC13, 40MPa, 1100x1100, B1F"),
    ("RC-Column_포디움(B,C구간)", 633600, "Column_RC, DC13, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633492, "Column_RC, DC14, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633472, "Column_RC, DC14, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 633490, "Column_RC, DC14, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633498, "Column_RC, DC14A, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633494, "Column_RC, DC14A, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 633496, "Column_RC, DC14A, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633550, "Column_RC, DC14B-1, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633530, "Column_RC, DC14B-1, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 633540, "Column_RC, DC14B-1, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633616, "Column_RC, DC14B, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633604, "Column_RC, DC14B, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 633606, "Column_RC, DC14B, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633622, "Column_RC, DC14C, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633618, "Column_RC, DC14C, 40MPa, 1000x1000, B1F"),
    ("RC-Column_포디움(B,C구간)", 633620, "Column_RC, DC14C, 40MPa, 900x900, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633526, "Column_RC, DC15, 40MPa, 800x800, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 4952499, "Column_RC, DC15, 40MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633516, "Column_RC, DC15, 40MPa, 800x800, B1F"),
    ("RC-Column_포디움(B,C구간)", 633127, "Column_RC, DC1A, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633117, "Column_RC, DC1A, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633107, "Column_RC, DC1A, 40MPa, 1300x1300, B1F"),
    ("RC-Column_포디움(B,C구간)", 633149, "Column_RC, DC1B, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633139, "Column_RC, DC1B, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633129, "Column_RC, DC1B, 40MPa, 1300x1300, B1F"),
    ("RC-Column_포디움(B,C구간)", 633155, "Column_RC, DC2, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633153, "Column_RC, DC2, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633151, "Column_RC, DC2, 40MPa, 1300x1300, B1F"),
    ("RC-Column_포디움(B,C구간)", 633161, "Column_RC, DC2A, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633159, "Column_RC, DC2A, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633157, "Column_RC, DC2A, 40MPa, 1300x1300, B1F"),
    ("RC-Column_포디움(B,C구간)", 633165, "Column_RC, DC3, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633163, "Column_RC, DC3, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 632957, "Column_RC, DC3, 40MPa, 1300x1300, B1F"),
    ("RC-Column_포디움(B,C구간)", 633171, "Column_RC, DC3A, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633169, "Column_RC, DC3A, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633167, "Column_RC, DC3A, 40MPa, 1300x1300, B1F"),
    ("RC-Column_포디움(B,C구간)", 633209, "Column_RC, DC5, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633199, "Column_RC, DC5, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633189, "Column_RC, DC5, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 633239, "Column_RC, DC5A, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633229, "Column_RC, DC5A, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633219, "Column_RC, DC5A, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 633269, "Column_RC, DC5B, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633259, "Column_RC, DC5B, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633249, "Column_RC, DC5B, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 633283, "Column_RC, DC6, 35MPa, 1400x1400, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633273, "Column_RC, DC6, 40MPa, 1500x1500, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633271, "Column_RC, DC6, 40MPa, 1600x1600, B1F"),
    ("RC-Column_포디움(B,C구간)", 633321, "Column_RC, DC6A, 35MPa, 1400x1400, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633311, "Column_RC, DC6A, 40MPa, 1400x1400, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633301, "Column_RC, DC6A, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 2267831, "Column_RC, DC7, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633333, "Column_RC, DC7, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633331, "Column_RC, DC7, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 633365, "Column_RC, DC7A, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633375, "Column_RC, DC7A, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633353, "Column_RC, DC7A, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 2360624, "Column_RC, DC7B, 35MPa, 1300x1300, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633396, "Column_RC, DC7B, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633394, "Column_RC, DC7B, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 633412, "Column_RC, DC8, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633410, "Column_RC, DC8, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633408, "Column_RC, DC8, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 633426, "Column_RC, DC8A, 35MPa, 1200x1200, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 633424, "Column_RC, DC8A, 40MPa, 1300x1300, 1F-2F"),
    ("RC-Column_포디움(B,C구간)", 633422, "Column_RC, DC8A, 40MPa, 1400x1400, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486494, "Column_RC, TC1, 35MPa, 800x800, 3F -4F"),
    ("RC-Column_포디움(B,C구간)", 486492, "Column_RC, TC1, 40MPa, 1200x1200, 1F"),
    ("RC-Column_포디움(B,C구간)", 486490, "Column_RC, TC1, 40MPa, 1400x1400, B2F-B1F"),
    ("RC-Column_포디움(B,C구간)", 486500, "Column_RC, TC1B, 35MPa, 800x800, 3F -4F"),
    ("RC-Column_포디움(B,C구간)", 788956, "Column_RC, TC1B, 40MPa, 1400x1400, 1F"),
    ("RC-Column_포디움(B,C구간)", 495444, "Column_RC, TC1B, 40MPa, 1400x1400, B2F -B1F"),
    ("RC-Column_포디움(B,C구간)", 486512, "Column_RC, TC2-1, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486508, "Column_RC, TC2, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 486506, "Column_RC, TC2, 40MPa, 1600x1200, B21F-1F"),
    ("RC-Column_포디움(B,C구간)", 486514, "Column_RC, TC31, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486516, "Column_RC, TC31B, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486520, "Column_RC, TC31C-1, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486518, "Column_RC, TC31C, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 486524, "Column_RC, TC31D, 35MPa, 800x800, 3F-4F"),
    ("RC-Column_포디움(B,C구간)", 941369, "Column_RC, TC31D, 40MPa, 1400x1400, B1F"),
    ("RC-Column_포디움(B,C구간)", 486526, "Column_RC, TC32, 35MPa, 800x800, 3F"),
    ("RC-Column_포디움(B,C구간)", 4568121, "Pedestal, NA, 30MPa, 500x500, 도면누락"),

    # ─── SRC-Column — 144 ─────────────────────────────────────────────
    ("SRC-Column", 1161264, "Column_SRC, AC1-1, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161262, "Column_SRC, AC1, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161292, "Column_SRC, AC10, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161266, "Column_SRC, AC1A, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161268, "Column_SRC, AC1B, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161270, "Column_SRC, AC2, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161274, "Column_SRC, AC3-1, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161272, "Column_SRC, AC3, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161276, "Column_SRC, AC3A, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161280, "Column_SRC, AC6-1, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161278, "Column_SRC, AC6, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161282, "Column_SRC, AC6A, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161286, "Column_SRC, AC7-1, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161284, "Column_SRC, AC7, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161290, "Column_SRC, AC8-1, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161288, "Column_SRC, AC8, 35MPa, 800x800, 5F-6F"),
    ("SRC-Column", 1161380, "Column_SRC, BC12, 40MPa, 1300x1300, 2F"),
    ("SRC-Column", 1161382, "Column_SRC, BC13, 40MPa, 1300x1300, 2F"),
    ("SRC-Column", 1161358, "Column_SRC, C131, 40MPa, 2000x1400, 2F"),
    ("SRC-Column", 1161306, "Column_SRC, C21-1, 35MPa, 800x900, 3F"),
    ("SRC-Column", 1161298, "Column_SRC, C21A, 35MPa, 800x800,  3F"),
    ("SRC-Column", 1161296, "Column_SRC, C21A, 40MPa, 900x900, 1F-2F"),
    ("SRC-Column", 2451927, "Column_SRC, C22-1, 35MPa, 800x800, 3F-4F"),
    ("SRC-Column", 1331370, "Column_SRC, C22-1, 40MPa, 900x900, 1F-2F"),
    ("SRC-Column", 1161312, "Column_SRC, C22-2, 35MPa, 800x800,  4F"),
    ("SRC-Column", 1161310, "Column_SRC, C22-2, 35MPa, 800x900,  3F"),
    ("SRC-Column", 1161308, "Column_SRC, C22-2, 40MPa, 900x900, 1F-2F"),
    ("SRC-Column", 1161302, "Column_SRC, C22, 35MPa, 800x800,  3F"),
    ("SRC-Column", 1161300, "Column_SRC, C22, 40MPa, 900x900, 1F-2F"),
    ("SRC-Column", 1161314, "Column_SRC, C22B, 35MPa, 800x800,  4F"),
    ("SRC-Column", 1161320, "Column_SRC, C22C-1, 35MPa, 800x800, 4F"),
    ("SRC-Column", 1161318, "Column_SRC, C22C-1, 35MPa, 800x900, 3F"),
    ("SRC-Column", 1161322, "Column_SRC, C23, 40MPa, 900x900, 2F"),
    ("SRC-Column", 1161326, "Column_SRC, C24, 40MPa, 900x900, 2F"),
    ("SRC-Column", 1161332, "Column_SRC, C24B, 35MPa, 800x800, 3F"),
    ("SRC-Column", 1161334, "Column_SRC, C25, 35MPa, 800x800, 4F"),
    ("SRC-Column", 1161338, "Column_SRC, C31-1, 40MPa, 900x900, 2F"),
    ("SRC-Column", 1161336, "Column_SRC, C31, 40MPa, 900x900, 2F"),
    ("SRC-Column", 1161340, "Column_SRC, C31A, 40MPa, 900x900, 1F-2F"),
    ("SRC-Column", 1161342, "Column_SRC, C32, 40MPa, 900x900, 2F"),
    ("SRC-Column", 1161344, "Column_SRC, C32B-1, 40MPa, 900x900, 1F-2F"),
    ("SRC-Column", 1161346, "Column_SRC, C32C, 40MPa, 900x900, 1F-2F"),
    ("SRC-Column", 1161348, "Column_SRC, C35, 35MPa, 800x800, 4F"),
    ("SRC-Column", 1161350, "Column_SRC, C36, 40MPa, 900x900, 2F"),
    ("SRC-Column", 1161352, "Column_SRC, C36A, 40MPa, 800x800, 2F"),
    ("SRC-Column", 1161356, "Column_SRC, C37, 40MPa, 900x900, 1F-2F"),
    ("SRC-Column", 1161394, "Column_SRC, CC15B, 35MPa, 800x800, 4F"),
    ("SRC-Column", 1161384, "Column_SRC, CC5A, 35MPa, 800x800, 4F"),
    ("SRC-Column", 1161402, "Column_SRC, DC1, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1616357, "Column_SRC, DC1, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161406, "Column_SRC, DC1, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161400, "Column_SRC, DC1, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161398, "Column_SRC, DC1, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161564, "Column_SRC, DC11A, 35MPa, 800x800, 6F"),
    ("SRC-Column", 1161566, "Column_SRC, DC12, 35MPa, 800x800, 6F"),
    ("SRC-Column", 1161568, "Column_SRC, DC14, 35MPa, 800x800, 6F"),
    ("SRC-Column", 1161570, "Column_SRC, DC14A, 35MPa, 800x800, 6F"),
    ("SRC-Column", 1161572, "Column_SRC, DC15, 35MPa, 800x800, 6F"),
    ("SRC-Column", 1586649, "Column_SRC, DC1A, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161408, "Column_SRC, DC1A, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161410, "Column_SRC, DC1A, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161404, "Column_SRC, DC1A, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161574, "Column_SRC, DC1A, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161416, "Column_SRC, DC1B, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161418, "Column_SRC, DC1B, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161420, "Column_SRC, DC1B, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161414, "Column_SRC, DC1B, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161412, "Column_SRC, DC1B, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161426, "Column_SRC, DC2, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161428, "Column_SRC, DC2, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161430, "Column_SRC, DC2, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161424, "Column_SRC, DC2, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161422, "Column_SRC, DC2, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161436, "Column_SRC, DC2A, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161438, "Column_SRC, DC2A, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161440, "Column_SRC, DC2A, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161434, "Column_SRC, DC2A, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161432, "Column_SRC, DC2A, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161446, "Column_SRC, DC3, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161448, "Column_SRC, DC3, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161450, "Column_SRC, DC3, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161444, "Column_SRC, DC3, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161442, "Column_SRC, DC3, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161456, "Column_SRC, DC3A, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161458, "Column_SRC, DC3A, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161460, "Column_SRC, DC3A, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161454, "Column_SRC, DC3A, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161452, "Column_SRC, DC3A, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161466, "Column_SRC, DC5, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161468, "Column_SRC, DC5, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161470, "Column_SRC, DC5, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161464, "Column_SRC, DC5, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161462, "Column_SRC, DC5, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161476, "Column_SRC, DC5A, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161478, "Column_SRC, DC5A, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161480, "Column_SRC, DC5A, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161474, "Column_SRC, DC5A, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161472, "Column_SRC, DC5A, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161486, "Column_SRC, DC5B, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161488, "Column_SRC, DC5B, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161490, "Column_SRC, DC5B, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161484, "Column_SRC, DC5B, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161482, "Column_SRC, DC5B, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161496, "Column_SRC, DC6, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161498, "Column_SRC, DC6, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161500, "Column_SRC, DC6, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161494, "Column_SRC, DC6, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161492, "Column_SRC, DC6, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161508, "Column_SRC, DC6A, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161510, "Column_SRC, DC6A, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161512, "Column_SRC, DC6A, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161504, "Column_SRC, DC6A, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161502, "Column_SRC, DC6A, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161518, "Column_SRC, DC7, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161520, "Column_SRC, DC7, 30MPa, 800x800, 17F"),
    ("SRC-Column", 1161522, "Column_SRC, DC7, 30MPa, 800x800, 20F"),
    ("SRC-Column", 1161516, "Column_SRC, DC7, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161514, "Column_SRC, DC7, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161528, "Column_SRC, DC7A, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161530, "Column_SRC, DC7A, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161532, "Column_SRC, DC7A, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161526, "Column_SRC, DC7A, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161524, "Column_SRC, DC7A, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161538, "Column_SRC, DC7B, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161540, "Column_SRC, DC7B, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161542, "Column_SRC, DC7B, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161536, "Column_SRC, DC7B, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161534, "Column_SRC, DC7B, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161548, "Column_SRC, DC8, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161550, "Column_SRC, DC8, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161552, "Column_SRC, DC8, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161546, "Column_SRC, DC8, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161544, "Column_SRC, DC8, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161558, "Column_SRC, DC8A, 30MPa, 800x800, 11F-13F"),
    ("SRC-Column", 1161560, "Column_SRC, DC8A, 30MPa, 800x800, 14F-17F"),
    ("SRC-Column", 1161562, "Column_SRC, DC8A, 30MPa, 800x800, 18F-20F"),
    ("SRC-Column", 1161556, "Column_SRC, DC8A, 30MPa, 900x900, 8F-10F"),
    ("SRC-Column", 1161554, "Column_SRC, DC8A, 35MPa, 1000x1000, 5F-7F"),
    ("SRC-Column", 1161362, "Column_SRC, TC1, 40MPa, 1200x1200, 2F"),
    ("SRC-Column", 1161368, "Column_SRC, TC2, 40MPa, 1200x1200, 2F"),
    ("SRC-Column", 1161372, "Column_SRC, TC31, 40MPa, 1200x1200, 2F"),
    ("SRC-Column", 1161374, "Column_SRC, TC31B, 40MPa, 1200x1200, 1F-2F"),
    ("SRC-Column", 1161376, "Column_SRC, TC31D, 40MPa, 1200x1200, 2F"),
    ("SRC-Column", 1161378, "Column_SRC, TC32, 40MPa, 1200x1200, 2F"),

    # ─── SRC_Circular-Column — 10 ─────────────────────────────────────
    ("SRC_Circular-Column", 1249509, "Column_SRC, BC12A, 40MPa, D1400, 2F"),
    ("SRC_Circular-Column", 1249511, "Column_SRC, BC12B, 40MPa, D1400, 2F"),
    ("SRC_Circular-Column", 1249497, "Column_SRC, C22A, 40MPa, D900, 1F"),
    ("SRC_Circular-Column", 1827058, "Column_SRC, C23c, 40MPa, D1400, 2F"),
    ("SRC_Circular-Column", 1827130, "Column_SRC, CC12A-1, 40MPa, D1400, 2F"),
    ("SRC_Circular-Column", 1249519, "Column_SRC, CC12A-3, 40MPa, D1400, 2F"),
    ("SRC_Circular-Column", 1249515, "Column_SRC, CC12A, 40MPa, D1400, 2F"),
    ("SRC_Circular-Column", 1249517, "Column_SRC, CC12B, 40MPa, D1400, 2F"),
    ("SRC_Circular-Column", 1249503, "Column_SRC, TC1B, 40MPa, D1400, 2F"),
    ("SRC_Circular-Column", 1249505, "Column_SRC, TC31C, 40MPa, D1400, 2F"),
]


def parse_type_name(name: str):
    """Return (재료, 부재마크, 강도, 단면, 적용층, 폭, 깊이/직경, 형상)."""
    parts = [p.strip() for p in name.split(",")]
    if len(parts) < 5:
        return ("", "", "", "", name, None, None, "")
    head = parts[0]                     # "Column_RC" / "Column_SRC" / "Pedestal"
    mark = parts[1]
    strength = parts[2]
    section = parts[3].replace(" ", "")
    level = parts[4]
    재료 = head.replace("Column_", "") if head.startswith("Column_") else head

    width = depth = None
    shape = ""
    if section.upper().startswith("D"):
        shape = "원형"
        m = re.match(r"D(\d+)", section, re.IGNORECASE)
        if m:
            width = depth = int(m.group(1))
    else:
        shape = "사각형"
        m = re.match(r"(\d+)x(\d+)", section, re.IGNORECASE)
        if m:
            width, depth = int(m.group(1)), int(m.group(2))

    return (재료, mark, strength, section, level, width, depth, shape)


HEADER = [
    "패밀리", "타입ID", "재료", "부재마크", "강도",
    "단면(mm)", "폭(mm)", "깊이/직경(mm)", "형상", "적용층", "타입명(원본)",
]

THIN = Side(border_style="thin", color="CCCCCC")
HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_FONT = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(name="맑은 고딕", size=10)


def write_sheet(ws, rows):
    ws.append(HEADER)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.row_dimensions[1].height = 24

    for fam, tid, name in rows:
        재료, mark, strength, section, level, w, d, shape = parse_type_name(name)
        ws.append([fam, tid, 재료, mark, strength, section, w, d, shape, level, name])

    last_row = ws.max_row
    last_col = len(HEADER)
    for row in ws.iter_rows(min_row=2, max_row=last_row, max_col=last_col):
        for c in row:
            c.font = BODY_FONT
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            c.alignment = Alignment(vertical="center")

    widths = [26, 10, 7, 12, 10, 14, 9, 13, 8, 18, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


def write_summary(ws, all_rows):
    ws.append(["패밀리", "타입 수"])
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.row_dimensions[1].height = 24

    counts = {}
    order = []
    for fam, _, _ in all_rows:
        if fam not in counts:
            order.append(fam)
            counts[fam] = 0
        counts[fam] += 1

    start = 2
    for i, fam in enumerate(order):
        r = start + i
        ws.cell(row=r, column=1, value=fam)
        ws.cell(row=r, column=2, value=counts[fam])
    total_row = start + len(order)
    ws.cell(row=total_row, column=1, value="합계").font = Font(name="맑은 고딕", bold=True)
    ws.cell(row=total_row, column=2, value=f"=SUM(B{start}:B{total_row-1})").font = Font(
        name="맑은 고딕", bold=True
    )

    for row in ws.iter_rows(min_row=2, max_row=total_row, max_col=2):
        for c in row:
            if c.font.bold is None or not c.font.bold:
                c.font = BODY_FONT
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            c.alignment = Alignment(vertical="center", horizontal="left" if c.column == 1 else "right")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12


def main():
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("요약")
    write_summary(summary, DATA)

    all_ws = wb.create_sheet("전체")
    write_sheet(all_ws, DATA)

    families = []
    seen = set()
    for fam, _, _ in DATA:
        if fam not in seen:
            seen.add(fam)
            families.append(fam)

    sheet_name_map = {
        "RC-Column_포디움(A구간)": "RC_A구간",
        "RC-Column_포디움(B,C구간)": "RC_BC구간",
        "SRC-Column": "SRC",
        "SRC_Circular-Column": "SRC_원형",
    }
    for fam in families:
        ws = wb.create_sheet(sheet_name_map.get(fam, fam[:31]))
        write_sheet(ws, [r for r in DATA if r[0] == fam])

    wb.save(OUTPUT)
    print(f"WROTE {OUTPUT}  rows={len(DATA)}")


if __name__ == "__main__":
    main()
