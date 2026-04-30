"""
Build RC_Column_Schedule.xlsx from parsed CAD column schedule data.
Mirrors the structure of RC_Beam_Schedule.xlsx for consistency.
Today (2026-04-30) we extracted only RC기둥 일람표-1 (1 of 13 sheets).
"""
import openpyxl
import io
import sys
from openpyxl.styles import Font, PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

OUT = r"C:\Users\ms.kim\Desktop\revit-mcp-v2\작업자료\2026-04-30\RC_Column_Schedule.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)

title_font = Font(bold=True, size=14, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="2D5BBE")
header_font = Font(bold=True, size=11)
header_fill = PatternFill("solid", fgColor="D9E5F5")
section_fill = PatternFill("solid", fgColor="F2F8E5")
warn_fill = PatternFill("solid", fgColor="FFF4D4")


# =========================================================================
# Sheet 1: 처리내역
# =========================================================================
ws = wb.create_sheet("처리내역")
ws["A1"] = "RC 기둥 일람표 → Revit 적용 처리 내역"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws.merge_cells("A1:C1")

rows = [
    None,
    ("단계", "결과", "비고"),
    (
        "1. CAD 일람표 추출 (RC기둥 일람표-1)",
        "4 column types 파싱 완료",
        "AutoCAD 도면 선택 영역에서 좌동(同) 처리 + 셀 위치 기반 그루핑",
    ),
    None,
    ("=== 다음 단계 (대기) ===", None, None),
    (
        "2. Revit 타입 신규 생성",
        "(예정) 6 types — VOID/철근 변경으로 분리",
        "기존 명명규칙: Column_RC, [부호], [강도], [B×D], [층범위]",
    ),
    (
        "3. 모델러 검토 필요 항목",
        "(대기)",
        "① 강도 (현재 가정: 40MPa)  ② C0B-1 VOID 범위 (B1F만? 1F만? 둘 다?)  ③ 부호 prefix (C0 vs BC0)",
    ),
    (
        "4. 인스턴스 배치",
        "(대기) — 도면 보면서 모델러 작업",
        "신규 type은 placeholder로 생성, 인스턴스는 별도 배치",
    ),
    None,
    ("=== 추출 데이터 요약 ===", None, None),
    ("선택된 entities", "429", "AutoCAD 선택 영역 (전체 도면 일부)"),
    ("텍스트 entities (MText/DBText)", "104", "데이터 셀, 헤더, 라벨 모두 포함"),
    ("발견된 column 부호", "4종 (C0, C0A, C0B-1, C0B-2)", None),
    ("커버 floor", "B2F ~ 4F~5F", None),
    None,
    ("=== 좌동 처리 로직 ===", None, None),
    ("정의", "좌동 = 같은 행에서 왼쪽 셀 값을 상속", None),
    (
        "패턴",
        "B2F만 데이터 입력 → B1F, 1F는 좌동 → B2F 값 상속",
        None,
    ),
    (
        "주철근 변경점 인식",
        "2F부터 22-D22 → 22-D19 변경 (C0B-1, C0B-2)",
        "사이즈는 같지만 철근 다름 → type 분리",
    ),
]

current = 2
for r in rows:
    if r is None:
        current += 1
        continue
    is_section = r[0] and (r[0].startswith("===") or r[0] == "단계")
    for col, val in enumerate(r, 1):
        if val is None:
            continue
        cell = ws.cell(row=current, column=col, value=val)
        if is_section:
            cell.font = header_font
            cell.fill = section_fill if r[0].startswith("===") else header_fill
    current += 1

ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 60


# =========================================================================
# Sheet 2: 타입정리 — 6 derived Revit types
# =========================================================================
ws = wb.create_sheet("타입정리")
ws["A1"] = "Revit 신규 생성 대상 (Column_RC 명명규칙)"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws.merge_cells("A1:C1")

for col, h in enumerate(
    ["타입 (분류, 부호, 강도, B×D, 층범위)", "Revit Type ID", "비고"], 1
):
    c = ws.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill

types = [
    (
        "Column_RC, C0, 40MPa, 700x700, B2F-1F",
        None,
        "⚠️ Revit 미반영 — 16-D22, D13@150/300, TIE BAR 2-D13",
    ),
    (
        "Column_RC, C0A, 40MPa, 800x800, B2F-B1F",
        None,
        "⚠️ Revit 미반영 — 20-D22, D13@150/300, TIE BAR 4-D13",
    ),
    (
        "Column_RC, C0B-1, 40MPa, 800x800, B2F",
        None,
        "⚠️ Revit 미반영 — 22-D22 (VOID 전 구간)",
    ),
    (
        "Column_RC, C0B-1, 40MPa, 800x800, 2F-5F",
        None,
        "⚠️ Revit 미반영 — 22-D19 (4F~5F 통합), VOID 후 구간",
    ),
    (
        "Column_RC, C0B-2, 40MPa, 800x800, B2F-1F",
        None,
        "⚠️ Revit 미반영 — 22-D22 (큰 철근 구간)",
    ),
    (
        "Column_RC, C0B-2, 40MPa, 800x800, 2F-4F",
        None,
        "⚠️ Revit 미반영 — 22-D19 (작은 철근 구간)",
    ),
]
for i, (name, tid, note) in enumerate(types, start=3):
    ws.cell(row=i, column=1, value=name)
    ws.cell(row=i, column=2, value=tid)
    cell = ws.cell(row=i, column=3, value=note)
    cell.fill = warn_fill

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 60


# =========================================================================
# Sheet 3: Revit매칭 (placeholder)
# =========================================================================
ws = wb.create_sheet("Revit매칭")
ws["A1"] = "Revit 현재 상태 vs 일람표 비교 (작업 시작 시 채움)"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws.merge_cells("A1:D1")
ws.cell(row=3, column=1, value="현재 상태: 일람표 추출만 완료, Revit cross-check 미수행")
ws.cell(
    row=4, column=1, value="다음 단계에서 Column 카테고리의 Column_RC 타입과 비교 예정"
)
ws.column_dimensions["A"].width = 80


# =========================================================================
# Sheet 4: RC Column Schedule — 좌동 풀어쓴 원본
# =========================================================================
ws = wb.create_sheet("RC Column Schedule")
ws["A1"] = "원본 일람표 (좌동 처리 적용 — 17 entry)"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws.merge_cells("A1:K1")

headers = [
    "#",
    "시트",
    "부호",
    "강도",
    "사이즈",
    "층",
    "주철근",
    "띠근 중앙부",
    "띠근 상하부",
    "TIE BAR",
    "비고",
]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill

data = [
    # C0 — 700x700
    (1, "RC기둥 일람표-1", "C0", "40MPa", "700x700", "B2F", "16-D22", "D13@300", "D13@150", "2-D13", "(원본)"),
    (2, "RC기둥 일람표-1", "C0", "40MPa", "700x700", "B1F", "16-D22", "D13@300", "D13@150", "2-D13", "좌동→B2F"),
    (3, "RC기둥 일람표-1", "C0", "40MPa", "700x700", "1F", "16-D22", "D13@300", "D13@150", "2-D13", "좌동→B1F"),
    # C0A — 800x800, B2F→B1F
    (4, "RC기둥 일람표-1", "C0A", "40MPa", "800x800", "B2F", "20-D22", "D13@300", "D13@150", "4-D13", "(원본)"),
    (5, "RC기둥 일람표-1", "C0A", "40MPa", "800x800", "B1F", "20-D22", "D13@300", "D13@150", "4-D13", "좌동→B2F"),
    # C0B-1 — VOID 구간 + 철근 변경
    (6, "RC기둥 일람표-1", "C0B-1", "40MPa", "800x800", "B2F", "22-D22", "D13@300", "D13@150", "4-D13", "(원본, 첫 type)"),
    (7, "RC기둥 일람표-1", "C0B-1", "40MPa", "VOID", "B1F", "VOID", "VOID", "VOID", "VOID", "VOID 구간 (Agent 보수적 해석)"),
    (8, "RC기둥 일람표-1", "C0B-1", "40MPa", "VOID", "1F", "VOID", "VOID", "VOID", "VOID", "VOID 구간 (Agent 보수적 해석)"),
    (9, "RC기둥 일람표-1", "C0B-1", "40MPa", "800x800", "2F", "22-D19", "D10@300", "D10@150", "4-D10", "(원본, 두 번째 type — 철근 변경)"),
    (10, "RC기둥 일람표-1", "C0B-1", "40MPa", "800x800", "3F", "22-D19", "D10@300", "D10@150", "4-D10", "좌동→2F"),
    (11, "RC기둥 일람표-1", "C0B-1", "40MPa", "800x800", "4F~5F", "22-D19", "D10@300", "D10@150", "4-D10", "좌동→3F"),
    # C0B-2 — 철근 변경 (B2F-1F vs 2F-4F)
    (12, "RC기둥 일람표-1", "C0B-2", "40MPa", "800x800", "B2F", "22-D22", "D13@300", "D13@150", "4-D13", "(원본, 첫 type)"),
    (13, "RC기둥 일람표-1", "C0B-2", "40MPa", "800x800", "B1F", "22-D22", "D13@300", "D13@150", "4-D13", "좌동→B2F"),
    (14, "RC기둥 일람표-1", "C0B-2", "40MPa", "800x800", "1F", "22-D22", "D13@300", "D13@150", "4-D13", "좌동→B1F"),
    (15, "RC기둥 일람표-1", "C0B-2", "40MPa", "800x800", "2F", "22-D19", "D10@300", "D10@150", "4-D10", "(원본, 두 번째 type — 철근 변경)"),
    (16, "RC기둥 일람표-1", "C0B-2", "40MPa", "800x800", "3F", "22-D19", "D10@300", "D10@150", "4-D10", "좌동→2F"),
    (17, "RC기둥 일람표-1", "C0B-2", "40MPa", "800x800", "4F", "22-D19", "D10@300", "D10@150", "4-D10", "좌동→3F"),
]

for row_data in data:
    target_row = 2 + row_data[0]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=target_row, column=col, value=val)
        if "VOID" in str(val):
            cell.fill = warn_fill

widths = [4, 18, 8, 8, 10, 8, 12, 12, 12, 10, 32]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w


# =========================================================================
# Sheet 5: 시트별 요약
# =========================================================================
ws = wb.create_sheet("시트별 요약")
ws["A1"] = "도면 시트별 column 분포"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws.merge_cells("A1:D1")

for col, h in enumerate(["시트", "column 부호 수", "부호 목록", "비고"], 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill

ws.cell(row=3, column=1, value="RC기둥 일람표-1")
ws.cell(row=3, column=2, value=4)
ws.cell(row=3, column=3, value="C0, C0A, C0B-1, C0B-2")
ws.cell(row=3, column=4, value="현재 추출된 시트")

ws.cell(row=4, column=1, value="(나머지 ~12개 시트)")
ws.cell(row=4, column=2, value="?")
ws.cell(row=4, column=3, value="(미추출)")
ws.cell(row=4, column=4, value="필요 시 별도 추출")

for i, w in enumerate([22, 16, 28, 22], 1):
    ws.column_dimensions[chr(64 + i)].width = w


# =========================================================================
# Sheet 6: 출처
# =========================================================================
ws = wb.create_sheet("출처")
ws["A1"] = "원본 도면 메타데이터"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws.merge_cells("A1:B1")

meta = [
    (
        "DWG 파일 경로",
        r"C:\Users\ms.kim\Desktop\01_구조구획\01.CAD(전기,설비)\섹터\A71 기둥일람표\YI-Y01P1B00-ST-A71-000301~000313 섹터(B,C섹터) RC기둥 일람표-1~13.dwg",
    ),
    ("도면 번호", "YI-Y01P1B00-ST-A71-000301~000313"),
    ("섹터", "B, C 섹터"),
    ("총 시트 수", "13 (RC기둥 일람표-1 ~ 13)"),
    ("이번 추출 시트", "RC기둥 일람표-1"),
    ("추출 일시", "2026-04-30"),
    ("선택된 entities", "429 (선택 영역 내 모든 객체)"),
    ("MText/DBText 추출", "104"),
    ("AutoCAD 버전", "AutoCAD 25.0.0.0"),
    (
        "파싱 도구",
        "Claude Code Agent + AutoCAD MCP (mcp__cad__get_selection_texts)",
    ),
    (
        "좌동 처리",
        "셀 위치 + 인코딩 패턴(\\xc1\\xc2\\xb5\\xbf cp949) 기반 자동 인식",
    ),
    (
        "관련 작업",
        "RC_Beam_Schedule.xlsx와 동일한 워크플로우 (보 → 기둥 확장)",
    ),
]
for i, (k, v) in enumerate(meta, 2):
    c = ws.cell(row=i, column=1, value=k)
    c.font = header_font
    c.fill = header_fill
    ws.cell(row=i, column=2, value=v)

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 100


# =========================================================================
# Reorder sheets in beam-schedule order
# =========================================================================
order = ["처리내역", "타입정리", "Revit매칭", "RC Column Schedule", "시트별 요약", "출처"]
wb._sheets = [wb[name] for name in order]

wb.save(OUT)
print(f"✅ 저장: {OUT}")
print(f"   시트 6개: {', '.join(order)}")
print(f"   타입정리: 6 신규 type 등록 (모두 ⚠️ Revit 미반영)")
print(f"   Schedule: 17 entry (좌동 풀어쓰기 + VOID 표기)")
